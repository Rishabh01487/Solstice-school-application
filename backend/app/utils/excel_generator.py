"""
EduNexus School — Excel Report Generator
Uses openpyxl to generate styled .xlsx exports.
"""

import io
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Brand Colors ──
PURPLE = "7C4DFF"
TEAL = "00BFA5"
DARK_BG = "1A1A2E"
LIGHT_BG = "F8F9FA"
WHITE = "FFFFFF"


def _style_header_row(ws, row: int, col_count: int, color: str = PURPLE):
    """Apply branded header styling to a row."""
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    header_font = Font(name="Segoe UI", bold=True, color=WHITE, size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def _add_title_block(ws, title: str, subtitle: str = ""):
    """Add a title block at the top of the worksheet."""
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"🏫 EduNexus School — {title}"
    title_cell.font = Font(name="Segoe UI", bold=True, size=16, color=PURPLE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    if subtitle:
        ws.merge_cells("A2:H2")
        sub_cell = ws["A2"]
        sub_cell.value = subtitle
        sub_cell.font = Font(name="Segoe UI", size=10, color="666666")
        sub_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20
        return 4  # Data starts at row 4
    return 3  # Data starts at row 3


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        adjusted = min(max(max_len + 2, min_width), max_width)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = adjusted


# ══════════════════════════════════════════════════════════
#  ATTENDANCE REPORT EXCEL
# ══════════════════════════════════════════════════════════

def generate_attendance_excel(data: Dict[str, Any]) -> bytes:
    """
    Generate an attendance report Excel file.
    
    data expects:
      - title: str
      - start_date: str
      - end_date: str
      - section_name: str
      - rows: list of dicts with student_name, admission_no, present, absent, late, excused, total, percentage
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    subtitle = f"Period: {data.get('start_date', '')} to {data.get('end_date', '')} | Section: {data.get('section_name', '')}"
    start_row = _add_title_block(ws, data.get("title", "Attendance Report"), subtitle)

    # Headers
    headers = ["#", "Student Name", "Admission No", "Present", "Absent", "Late", "Excused", "Total", "Attendance %"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=header)
    _style_header_row(ws, start_row, len(headers), TEAL)

    # Data rows
    data_font = Font(name="Segoe UI", size=10)
    alt_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
    center_align = Alignment(horizontal="center")

    for idx, row in enumerate(data.get("rows", []), 1):
        r = start_row + idx
        ws.cell(row=r, column=1, value=idx).font = data_font
        ws.cell(row=r, column=2, value=row["student_name"]).font = Font(name="Segoe UI", bold=True, size=10)
        ws.cell(row=r, column=3, value=row["admission_no"]).font = data_font
        ws.cell(row=r, column=4, value=row["present"]).font = Font(name="Segoe UI", size=10, color="2E7D32")
        ws.cell(row=r, column=5, value=row["absent"]).font = Font(name="Segoe UI", size=10, color="C62828")
        ws.cell(row=r, column=6, value=row["late"]).font = Font(name="Segoe UI", size=10, color="E65100")
        ws.cell(row=r, column=7, value=row["excused"]).font = Font(name="Segoe UI", size=10, color="1565C0")
        ws.cell(row=r, column=8, value=row["total"]).font = data_font
        pct_cell = ws.cell(row=r, column=9, value=row["percentage"] / 100)
        pct_cell.number_format = "0.0%"
        pct_cell.font = Font(name="Segoe UI", bold=True, size=10)

        for col in range(1, 10):
            ws.cell(row=r, column=col).alignment = center_align
            if idx % 2 == 0:
                ws.cell(row=r, column=col).fill = alt_fill

    # Footer
    footer_row = start_row + len(data.get("rows", [])) + 2
    ws.cell(row=footer_row, column=1, value=f"Generated: {datetime.utcnow().strftime('%B %d, %Y')}").font = Font(
        name="Segoe UI", size=8, color="999999", italic=True
    )

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ══════════════════════════════════════════════════════════
#  STUDENT LIST EXCEL
# ══════════════════════════════════════════════════════════

def generate_student_list_excel(data: Dict[str, Any]) -> bytes:
    """
    Generate a student list Excel file.
    
    data expects:
      - title: str
      - students: list of dicts with name, admission_no, gender, status, email, enrollment_date
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Student List"

    start_row = _add_title_block(ws, data.get("title", "Student List"))

    headers = ["#", "Full Name", "Admission No", "Gender", "Status", "Email", "Enrollment Date"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=header)
    _style_header_row(ws, start_row, len(headers))

    data_font = Font(name="Segoe UI", size=10)
    alt_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")

    for idx, student in enumerate(data.get("students", []), 1):
        r = start_row + idx
        ws.cell(row=r, column=1, value=idx).font = data_font
        ws.cell(row=r, column=2, value=student["name"]).font = Font(name="Segoe UI", bold=True, size=10)
        ws.cell(row=r, column=3, value=student["admission_no"]).font = data_font
        ws.cell(row=r, column=4, value=student["gender"]).font = data_font
        ws.cell(row=r, column=5, value=student["status"]).font = data_font
        ws.cell(row=r, column=6, value=student["email"]).font = data_font
        ws.cell(row=r, column=7, value=student["enrollment_date"]).font = data_font

        if idx % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=r, column=col).fill = alt_fill

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ══════════════════════════════════════════════════════════
#  GRADES REPORT EXCEL
# ══════════════════════════════════════════════════════════

def generate_grades_excel(data: Dict[str, Any]) -> bytes:
    """
    Generate a grades report Excel file.
    
    data expects:
      - title: str
      - term_name: str
      - rows: list of dicts with student_name, admission_no, subjects (list of {name, score, grade})
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades Report"

    subtitle = f"Term: {data.get('term_name', '')}"
    start_row = _add_title_block(ws, data.get("title", "Grades Report"), subtitle)

    # Determine subject columns from first row
    subject_names = []
    if data.get("rows"):
        subject_names = [s["name"] for s in data["rows"][0].get("subjects", [])]

    headers = ["#", "Student Name", "Admission No"] + subject_names + ["Average", "Grade"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=header)
    _style_header_row(ws, start_row, len(headers))

    data_font = Font(name="Segoe UI", size=10)
    center_align = Alignment(horizontal="center")

    for idx, row in enumerate(data.get("rows", []), 1):
        r = start_row + idx
        ws.cell(row=r, column=1, value=idx).font = data_font
        ws.cell(row=r, column=2, value=row["student_name"]).font = Font(name="Segoe UI", bold=True, size=10)
        ws.cell(row=r, column=3, value=row["admission_no"]).font = data_font

        scores = []
        for s_idx, subj in enumerate(row.get("subjects", []), 4):
            score = subj.get("score", 0)
            cell = ws.cell(row=r, column=s_idx, value=score)
            cell.font = data_font
            cell.alignment = center_align
            scores.append(score)

        base_col = 4 + len(subject_names)
        avg = sum(scores) / len(scores) if scores else 0
        ws.cell(row=r, column=base_col, value=round(avg, 1)).font = Font(name="Segoe UI", bold=True, size=10)
        ws.cell(row=r, column=base_col, value=round(avg, 1)).alignment = center_align

        grade = row.get("overall_grade", "")
        ws.cell(row=r, column=base_col + 1, value=grade).font = Font(name="Segoe UI", bold=True, size=10)

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
